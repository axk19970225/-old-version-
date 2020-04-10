import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from sklearn.svm import SVC
from sklearn.metrics import confusion_matrix
from sklearn.metrics import precision_score, recall_score
from sklearn.externals import joblib
import joblib

#画precision/recall曲线函数
def plot_precision_recall_threshold(precisions,recalls,thresholds):
    plt.plot(thresholds,precisions[:-1],"b--",label="Precision")
    plt.plot(thresholds,recalls[:-1], "g--", label="Recall")
    plt.xlabel("Threshold")
    plt.ylim([0,1])

#画ROC曲线函数
def plot_roc_curve(fpr,tpr,label=None):
    plt.plot(fpr,tpr,linewidth = 2,label = label)
    plt.plot([0,1],[0,1],'k--')
    plt.axis([0,1,0,1])
    plt.xlabel("Flase Positive Rate")
    plt.ylabel("True Positive Rate")




TrainPositive = np.load(r"/home/axk/chipsvm/TrainPositive.npy")
TrainNegative = np.load(r"/home/axk/chipsvm/TrainNegative.npy")
TestPositive = np.load(r"/home/axk/chipsvm/TestPositive.npy")
TestNegative = np.load(r"/home/axk/chipsvm/TestNegative.npy")

np.random.seed(42)
PositiveTrainIndex = np.random.choice(np.arange(len(TrainPositive)),45000)
PositiveTrain = []
for i in PositiveTrainIndex:
    PositiveTrain.append(TrainPositive[i])
PositiveTestIndex = np.random.choice(np.arange(len(TestPositive)),500)
PositiveTest = []
for i in PositiveTestIndex:
    PositiveTest.append(TestPositive[i])

#Positive加标签
PositiveTrainLabel = []
for i in range(len(PositiveTrain)):
    PositiveTrainLabel.append(1)
PositiveTestLabel = []
for i in range(len(PositiveTest)):
    PositiveTestLabel.append(1)




NegativeTestIndex = np.random.choice(np.arange(len(TestNegative)),500)
NegativeTest = []
for i in NegativeTestIndex:
    NegativeTest.append(TestNegative[i])
NegativeTestLabel = []
for i in range(len(NegativeTest)):
    NegativeTestLabel.append(-1)


#扩充Negative数量,
NegativeTrain = TrainNegative
temp = TrainNegative
for i in range(9):
    NegativeTrain = np.vstack((NegativeTrain,temp)) # 45060个
NegativeTrainLabel = []
for i in range(len(NegativeTrain)):
    NegativeTrainLabel.append(-1)


#将训练数据合并，并打乱
TrainData_Temp = list(PositiveTrain) + list(NegativeTrain)
TrainLabel_Temp = PositiveTrainLabel + NegativeTrainLabel
TrainIndex = np.arange(len(TrainLabel_Temp ))
TrainIndex = np.random.permutation(TrainIndex)
TrainData = []
TrainLabel = []
for i in TrainIndex:
    TrainData.append(TrainData_Temp[i])
    TrainLabel.append(TrainLabel_Temp[i])


#将测试数据合并，并打乱
TestData_Temp = list(PositiveTest) + list(NegativeTest)
TestLabel_Temp = PositiveTestLabel + NegativeTestLabel
TestIndex = np.arange(len(TestLabel_Temp))
TestIndex = np.random.permutation(TestIndex)
TestData = []
TestLabel = []
for i in TestIndex:
    TestData.append(TestData_Temp[i])
    TestLabel.append(TestLabel_Temp[i])


print(np.array(PositiveTrain).shape)
print(np.array(NegativeTrain).shape)
print(np.array(PositiveTest).shape)
print(np.array(NegativeTest).shape)



print("数据处理完成")

start = 2
for C in [0.00001,0.0001,0.001,0.01,0.1,1,10,100,1000,10000,100000]:
    print("     ")
    print("     ")
    print("C = {}".format(C))
    print("     ")
    print("     ")
    #训练
    # clf = SVC(kernel="rbf", C = 1.0, gamma= "auto")
    clf = SVC(kernel="linear", C = C)
    clf.fit(TrainData, TrainLabel)
    Modalpath = r"/home/axk/linearsvm/SVMmodal/svmModellinear" + "c" + str(C) + r".plk"
    joblib.dump(clf, Modalpath)
    TestPredict = clf.predict(TestData)
    print("训练完成")
    #评价
    print("返回给定测试集和对应标签的平均准确率")
    accuracy = clf.score(TestData, TestLabel)
    print(accuracy)
    print("混淆矩阵：")
    conf = confusion_matrix(TestLabel,TestPredict)
    print(conf)
    print("precision : ")
    precision = precision_score(TestLabel,TestPredict)
    print(precision)
    print("recall : ")
    recall = recall_score(TestLabel,TestPredict)
    print(recall)
    print("FPR : ")
    FPR = conf[0,1]/(conf[0,1] + conf[0,0])
    print(conf[0,1]/(conf[0,1] + conf[0,0]))


    #交叉评估
    from sklearn.model_selection import cross_val_score
    score = cross_val_score(clf,TestData,TestLabel,cv=3,scoring="accuracy")
    print("cross_val_score ： {}".format(score))

    # precision-recall曲线
    label_score = clf.decision_function(TestData)
    from sklearn.metrics import precision_recall_curve

    precisions, recalls, thresholds = precision_recall_curve(TestLabel, label_score)
    prname = "pr" + str(gamma) + "vs" + str(C)
    plt.figure("{}".format(prname))
    plot_precision_recall_threshold(precisions, recalls, thresholds)
    prpath = r"/home/axk/linearsvm/prpicture/linear" + "c" + str(C) + r".png"
    plt.savefig(prpath)
    plt.close()

    # ROC曲线
    from sklearn.metrics import roc_curve

    fpr, tpr, thresholds = roc_curve(TestLabel, label_score)
    rocname = "roc" + str(gamma) + "vs" + str(C)
    plt.figure("{}".format(rocname))
    plot_roc_curve(fpr, tpr)
    rocpath = r"/home/axk/linearsvm/rocpicture/linear" + "c" + str(C) + r".png"
    plt.savefig(rocpath)
    plt.close()


    #AUC
    from sklearn.metrics import roc_auc_score
    print("AUC : ")
    AUC = roc_auc_score(TestLabel, label_score)
    print(AUC)

    #储存数据到excel
    wb = openpyxl.load_workbook(r"/home/axk/linearsvm/result.xlsx")
    ws = wb.active
    ws["B" + str(start)] = accuracy
    ws["C" + str(start)] = precision
    ws["D" + str(start)] = recall
    ws["E" + str(start)] = FPR
    ws["F" + str(start)] = conf[0,0]
    ws["G" + str(start)] = conf[0,1]
    ws["H" + str(start)] = conf[1,0]
    ws["I" + str(start)] = conf[1,1]
    ws["J" + str(start)] = AUC
    ws["K" + str(start)] = str(score)
    start += 1
    wb.save(r"/home/axk/linearsvm/result.xlsx")







